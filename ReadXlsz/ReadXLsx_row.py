import openpyxl

path = "/Users/aravindanathdm/PycharmProjects/PythonSelenium/testData.xlsx"


wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active


max_row = sheet_obj.max_column


for i in range(1,max_row+1):
   cell_obj = sheet_obj.cell(row=i,column=1)
   print(cell_obj.value)