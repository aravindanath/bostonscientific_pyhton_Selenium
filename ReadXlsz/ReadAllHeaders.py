import openpyxl

path = "/Users/aravindanathdm/PycharmProjects/PythonSelenium/testData.xlsx"


wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active

max_Col = sheet_obj.max_column


for i in range(1,max_Col+1):
   cell_obj = sheet_obj.cell(row=1,column=i)
   print(cell_obj.value)