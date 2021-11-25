import openpyxl

path = "/Users/aravindanathdm/PycharmProjects/PythonSelenium/testData.xlsx"


wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active

print(sheet_obj.max_column)
print(sheet_obj.max_row)